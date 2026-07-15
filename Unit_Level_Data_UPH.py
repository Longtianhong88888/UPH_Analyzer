import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os
import csv
from datetime import datetime
import tkinter as tk
from tkinter import filedialog

# ============================================================
# 配置参数
# ============================================================
LASER_INDEX_COLUMN = "LaserIndex"
PICK_TIME_COLUMN = "ModulePickTime"
CT_THRESHOLD_LOW = 15.0   # CT低于此值视为异常（标红）
CT_THRESHOLD_HIGH = 20.0  # CT高于此值视为异常（标红）


def format_seconds_to_time(seconds):
    """将秒数转换为 时:分:秒.毫秒 格式"""
    if seconds is None or seconds == "":
        return ""
    if not isinstance(seconds, (int, float)):
        return ""
    hours = int(seconds // 3600)
    minutes = int((seconds % 3600) // 60)
    secs = seconds % 60
    return f"{hours}:{minutes:02d}:{secs:06.3f}"


def clean_value_for_excel(value):
    """
    清理值，确保写入Excel的是纯数值而非公式
    """
    if value is None:
        return None
    if isinstance(value, str):
        # 如果字符串以 = 开头，说明是公式，尝试提取数值
        if value.startswith('='):
            try:
                cleaned = value.replace('=', '').strip()
                if any(op in cleaned for op in ['+', '-', '*', '/']):
                    try:
                        import re
                        if re.match(r'^[\d+\-*/.()\s]+$', cleaned):
                            result = eval(cleaned)
                            if isinstance(result, (int, float)):
                                return round(result, 6)
                    except:
                        pass
                try:
                    return float(cleaned)
                except ValueError:
                    pass
                return cleaned
            except:
                return value
        try:
            if value.replace('.', '', 1).replace('-', '', 1).isdigit():
                return float(value)
        except:
            pass
    elif isinstance(value, (int, float)):
        return value
    return value


def remove_all_formulas_from_workbook(wb):
    """
    遍历工作簿中所有Sheet的所有单元格，将公式替换为计算后的数值
    这是确保Numbers不报错的关键步骤
    """
    print("🔧 正在清除所有公式...")
    formula_count = 0
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                # 检查是否为公式
                if cell.data_type == 'f' or (isinstance(cell.value, str) and cell.value.startswith('=')):
                    try:
                        # 获取公式计算后的值
                        # openpyxl 在读取时如果使用 data_only=True，cell.value 已经是计算值
                        # 但为了安全，我们直接读取并重新赋值
                        if hasattr(cell.value, 'result'):
                            cell.value = cell.value.result
                        else:
                            # 尝试将公式字符串转换为数值
                            val_str = str(cell.value)
                            if val_str.startswith('='):
                                # 尝试提取数值
                                cleaned = val_str.replace('=', '').strip()
                                try:
                                    cell.value = float(cleaned)
                                except:
                                    cell.value = None
                            else:
                                cell.value = cell.value
                        formula_count += 1
                    except:
                        cell.value = None
    print(f"✅ 已清除 {formula_count} 个公式")
    return wb


def parse_time_to_datetime(value):
    """
    智能解析时间，支持多种格式
    """
    if value is None:
        return None
    
    if isinstance(value, datetime):
        return value
    
    if isinstance(value, (int, float)):
        try:
            from datetime import timedelta
            if value > 40000:
                return datetime(1899, 12, 30) + timedelta(days=value)
            else:
                return datetime(1900, 1, 1) + timedelta(seconds=value)
        except:
            return None
    
    if isinstance(value, str):
        value = value.strip()
        formats = [
            "%Y-%m-%d %H:%M:%S.%f",
            "%Y-%m-%d %H:%M:%S",
            "%Y/%m/%d %H:%M:%S.%f",
            "%Y/%m/%d %H:%M:%S",
            "%H:%M:%S.%f",
            "%H:%M:%S",
            "%Y-%m-%dT%H:%M:%S.%f",
            "%Y-%m-%dT%H:%M:%S",
            "%m/%d/%Y %H:%M:%S",
            "%m/%d/%Y %H:%M",
        ]
        for fmt in formats:
            try:
                return datetime.strptime(value, fmt)
            except ValueError:
                continue
        
        if '.' in value:
            try:
                base_time = value.split('.')[0]
                return datetime.strptime(base_time, "%Y-%m-%d %H:%M:%S")
            except:
                pass
        
        return None
    
    return None


def read_data_from_file(file_path):
    """
    读取数据，自动识别CSV或Excel格式
    """
    file_ext = os.path.splitext(file_path)[1].lower()
    
    if file_ext in ['.csv']:
        data = []
        headers = []
        try:
            for encoding in ['utf-8', 'gbk', 'gb2312', 'latin-1']:
                try:
                    with open(file_path, 'r', encoding=encoding) as f:
                        reader = csv.reader(f)
                        headers = next(reader)
                        for row in reader:
                            cleaned_row = [clean_value_for_excel(cell) for cell in row]
                            data.append(cleaned_row)
                    print(f"✅ 使用编码 {encoding} 成功读取CSV文件")
                    break
                except UnicodeDecodeError:
                    continue
                except StopIteration:
                    return None, None
            else:
                print("❌ 无法解码CSV文件，请检查文件编码")
                return None, None
        except Exception as e:
            print(f"❌ 读取CSV文件失败：{e}")
            return None, None
        
        headers = [h.strip().replace('\ufeff', '') for h in headers]
        return headers, data
    
    elif file_ext in ['.xlsx', '.xlsm', '.xltx', '.xltm']:
        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
            ws = wb.active
        except Exception as e:
            print(f"❌ 读取Excel文件失败：{e}")
            return None, None
        
        headers = []
        for col in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=1, column=col).value
            if cell_value:
                headers.append(str(cell_value).strip())
            else:
                headers.append(f"Column{col}")
        
        data = []
        for row in range(2, ws.max_row + 1):
            row_data = []
            for col in range(1, ws.max_column + 1):
                val = ws.cell(row=row, column=col).value
                row_data.append(clean_value_for_excel(val))
            data.append(row_data)
        
        wb.close()
        return headers, data
    
    else:
        print(f"❌ 不支持的文件格式：{file_ext}，请使用 .csv 或 .xlsx 文件")
        return None, None


def calculate_uph(input_file_path, output_file_path=None):
    """
    从机台日志中计算UPH，支持CSV和Excel格式
    """
    if output_file_path is None:
        base_name = os.path.splitext(input_file_path)[0]
        output_file_path = f"{base_name}_UPH分析.xlsx"
    
    # 1. 读取数据
    headers, data = read_data_from_file(input_file_path)
    if headers is None or data is None:
        return
    
    print(f"✅ 读取成功：{len(headers)} 列，{len(data)} 行数据")
    
    # 查找关键列
    try:
        laser_idx = headers.index(LASER_INDEX_COLUMN)
        pick_time_idx = headers.index(PICK_TIME_COLUMN)
    except ValueError as e:
        print(f"❌ 找不到必要的列：{e}")
        print(f"   当前表头：{headers}")
        return
    
    print(f"✅ 找到列：{LASER_INDEX_COLUMN}（列{laser_idx+1}），{PICK_TIME_COLUMN}（列{pick_time_idx+1}）")
    
    # 2. 按LaserIndex分类数据
    data_by_laser = {}
    
    for row_data in data:
        if row_data is None or len(row_data) <= laser_idx:
            continue
        laser_value = row_data[laser_idx]
        if laser_value is None:
            continue
        
        laser_key = str(laser_value).strip()
        if laser_key not in data_by_laser:
            data_by_laser[laser_key] = []
        data_by_laser[laser_key].append(row_data)
    
    if not data_by_laser:
        print("❌ 未找到任何有效数据")
        return
    
    print(f"✅ 共找到 {len(data_by_laser)} 个不同的 LaserIndex")
    for key in data_by_laser.keys():
        print(f"   - {key}: {len(data_by_laser[key])} 条记录")
    
    # 3. 创建输出工作簿
    wb_output = openpyxl.Workbook()
    default_sheet = wb_output.active
    wb_output.remove(default_sheet)
    
    # 定义样式
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    light_red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
    green_fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
    blue_fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
    header_font = Font(bold=True)
    center_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    
    # 存储每个LaserIndex的统计数据
    laser_stats = {}
    all_ct_values = []
    all_normal_avgs = []
    
    # 4. 处理每个LaserIndex
    for laser_key, rows in data_by_laser.items():
        sheet_name = f"{laser_key}"[:31]
        ws = wb_output.create_sheet(title=sheet_name)
        
        # 写入表头
        new_headers = []
        for i, h in enumerate(headers):
            new_headers.append(h)
            if i == laser_idx:
                new_headers.append("CT（秒）")
                new_headers.append("产品分类")
        
        for col_num, h in enumerate(new_headers, start=1):
            ws.cell(row=1, column=col_num, value=h)
        
        # 获取所有时间
        time_values = []
        for row_data in rows:
            time_val = row_data[pick_time_idx] if pick_time_idx < len(row_data) else None
            time_dt = parse_time_to_datetime(time_val)
            time_values.append(time_dt)
        
        # 计算CT
        normal_ct_list = []
        abnormal_ct_list = []
        ct_list = []
        
        for i in range(len(rows)):
            row_data = rows[i]
            
            # 计算CT
            ct_value = None
            if i > 0 and time_values[i] and time_values[i-1]:
                delta = (time_values[i] - time_values[i-1]).total_seconds()
                if delta >= 0:
                    ct_value = round(delta, 3)
                else:
                    delta = (time_values[i] - time_values[i-1]).total_seconds() + 86400
                    if delta >= 0:
                        ct_value = round(delta, 3)
            
            # 写入数据行
            col_idx = 1
            for j, val in enumerate(row_data):
                ws.cell(row=i+2, column=col_idx, value=clean_value_for_excel(val))
                col_idx += 1
                if j == laser_idx:
                    if i > 0 and ct_value is not None:
                        ws.cell(row=i+2, column=col_idx, value=ct_value)
                        ct_list.append(ct_value)
                        
                        if ct_value > CT_THRESHOLD_HIGH or ct_value < CT_THRESHOLD_LOW:
                            abnormal_ct_list.append(ct_value)
                            for col in range(1, len(new_headers) + 1):
                                cell = ws.cell(row=i+2, column=col)
                                cell.fill = light_red_fill
                            ws.cell(row=i+2, column=col_idx+1, value="异常").fill = red_fill
                        else:
                            normal_ct_list.append(ct_value)
                            ws.cell(row=i+2, column=col_idx+1, value="正常").fill = green_fill
                    else:
                        ws.cell(row=i+2, column=col_idx, value="")
                        ws.cell(row=i+2, column=col_idx+1, value="")
                    col_idx += 2
        
        # 存储统计
        laser_stats[laser_key] = {
            "normal": normal_ct_list,
            "abnormal": abnormal_ct_list,
            "all_ct": ct_list
        }
        all_ct_values.extend(ct_list)
        if normal_ct_list:
            all_normal_avgs.append(sum(normal_ct_list) / len(normal_ct_list))
        
        # 底部统计
        stat_row = len(rows) + 4
        ws.cell(row=stat_row, column=1, value="=== 统计信息 ===")
        ws.cell(row=stat_row, column=1).font = Font(bold=True, size=11)
        
        stat_labels = [
            ("正常产品数", len(normal_ct_list)),
            ("异常产品数", len(abnormal_ct_list)),
            ("正常产品最大CT", max(normal_ct_list) if normal_ct_list else ""),
            ("正常产品最小CT", min(normal_ct_list) if normal_ct_list else ""),
            ("正常产品平均CT", round(sum(normal_ct_list)/len(normal_ct_list), 3) if normal_ct_list else ""),
            ("异常产品最大CT", max(abnormal_ct_list) if abnormal_ct_list else ""),
            ("异常产品最小CT", min(abnormal_ct_list) if abnormal_ct_list else ""),
            ("异常产品平均CT", round(sum(abnormal_ct_list)/len(abnormal_ct_list), 3) if abnormal_ct_list else ""),
        ]
        
        for i, (label, value) in enumerate(stat_labels, start=1):
            row = stat_row + i
            ws.cell(row=row, column=1, value=label)
            ws.cell(row=row, column=2, value=value)
            ws.cell(row=row, column=1).font = header_font
        
        # 调整列宽
        for col in ws.columns:
            max_length = 0
            column_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        print(f"✅ Sheet '{sheet_name}' 已创建：正常 {len(normal_ct_list)} 个，异常 {len(abnormal_ct_list)} 个")
    
    # ============================================================
    # 5. UPH分析Summary
    # ============================================================
    ws_summary = wb_output.create_sheet(title="UPH分析", index=0)
    
    ws_summary.cell(row=1, column=1, value="UPH分析报告")
    ws_summary.cell(row=1, column=1).font = Font(bold=True, size=16)
    
    ws_summary.cell(row=3, column=1, value="LaserIndex")
    ws_summary.cell(row=3, column=2, value="产品总数")
    ws_summary.cell(row=3, column=3, value="正常产品数")
    ws_summary.cell(row=3, column=4, value="异常产品数")
    ws_summary.cell(row=3, column=5, value="正常平均CT(秒)")
    ws_summary.cell(row=3, column=6, value="异常平均CT(秒)")
    
    for col in range(1, 7):
        cell = ws_summary.cell(row=3, column=col)
        cell.font = header_font
        cell.alignment = center_alignment
        cell.border = thin_border
    
    row = 4
    total_normal_count = 0
    total_abnormal_count = 0
    total_products = 0
    
    for laser_key, stats in laser_stats.items():
        normal_list = stats["normal"]
        abnormal_list = stats["abnormal"]
        
        normal_count = len(normal_list)
        abnormal_count = len(abnormal_list)
        total_count = normal_count + abnormal_count
        
        normal_avg = round(sum(normal_list)/len(normal_list), 3) if normal_list else ""
        abnormal_avg = round(sum(abnormal_list)/len(abnormal_list), 3) if abnormal_list else ""
        
        ws_summary.cell(row=row, column=1, value=laser_key)
        ws_summary.cell(row=row, column=2, value=total_count)
        ws_summary.cell(row=row, column=3, value=normal_count)
        ws_summary.cell(row=row, column=4, value=abnormal_count)
        ws_summary.cell(row=row, column=5, value=normal_avg)
        ws_summary.cell(row=row, column=6, value=abnormal_avg)
        
        for col in range(1, 7):
            cell = ws_summary.cell(row=row, column=col)
            cell.alignment = center_alignment
            cell.border = thin_border
        
        total_normal_count += normal_count
        total_abnormal_count += abnormal_count
        total_products += total_count
        row += 1
    
    # 汇总行
    ws_summary.cell(row=row, column=1, value="=== 汇总 ===")
    ws_summary.cell(row=row, column=2, value=total_products)
    ws_summary.cell(row=row, column=3, value=total_normal_count)
    ws_summary.cell(row=row, column=4, value=total_abnormal_count)
    for col in range(1, 7):
        cell = ws_summary.cell(row=row, column=col)
        cell.font = Font(bold=True)
        cell.alignment = center_alignment
        cell.border = thin_border
        cell.fill = blue_fill
    
    row += 2
    
    # UPH计算
    sum_normal_avgs = sum(all_normal_avgs) if all_normal_avgs else 0
    laser_count = len(all_normal_avgs) if all_normal_avgs else 1
    
    if sum_normal_avgs > 0:
        normal_avg_ct = sum_normal_avgs / laser_count
        normal_uph = round((3600 / normal_avg_ct) * 8, 2)
    else:
        normal_uph = 0
    
    if all_ct_values:
        total_ct_sum = sum(all_ct_values)
        total_ct_count = len(all_ct_values)
        actual_avg_ct = total_ct_sum / total_ct_count
        actual_uph = round((3600 / actual_avg_ct) * 8, 2)
    else:
        actual_uph = 0
    
    ws_summary.cell(row=row, column=1, value="UPH计算结果")
    ws_summary.cell(row=row, column=1).font = Font(bold=True, size=12)
    
    row += 1
    ws_summary.cell(row=row, column=1, value="所有LaserIndex正常产品平均值之和")
    ws_summary.cell(row=row, column=2, value=round(sum_normal_avgs, 3))
    ws_summary.cell(row=row, column=1).font = header_font
    
    row += 1
    ws_summary.cell(row=row, column=1, value="LaserIndex数量")
    ws_summary.cell(row=row, column=2, value=laser_count)
    ws_summary.cell(row=row, column=1).font = header_font
    
    row += 1
    ws_summary.cell(row=row, column=1, value="所有产品CT求和")
    ws_summary.cell(row=row, column=2, value=round(sum(all_ct_values), 3) if all_ct_values else 0)
    ws_summary.cell(row=row, column=1).font = header_font
    
    row += 1
    ws_summary.cell(row=row, column=1, value="所有产品总数")
    ws_summary.cell(row=row, column=2, value=len(all_ct_values))
    ws_summary.cell(row=row, column=1).font = header_font
    
    row += 2
    
    ws_summary.cell(row=row, column=1, value="正常产品UPH预估")
    ws_summary.cell(row=row, column=2, value=normal_uph)
    for col in range(1, 3):
        cell = ws_summary.cell(row=row, column=col)
        cell.font = Font(bold=True, size=12)
        cell.alignment = center_alignment
        cell.fill = green_fill
        cell.border = thin_border
    
    row += 1
    ws_summary.cell(row=row, column=1, value="实际UPH")
    ws_summary.cell(row=row, column=2, value=actual_uph)
    for col in range(1, 3):
        cell = ws_summary.cell(row=row, column=col)
        cell.font = Font(bold=True, size=12)
        cell.alignment = center_alignment
        cell.fill = green_fill
        cell.border = thin_border
    
    # 调整列宽
    for col in ws_summary.columns:
        max_length = 0
        column_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = min(max_length + 2, 35)
        ws_summary.column_dimensions[column_letter].width = adjusted_width
    
    # ============================================================
    # 【关键】保存前清除所有公式
    # ============================================================
    wb_output = remove_all_formulas_from_workbook(wb_output)
    
    # 保存
    wb_output.save(output_file_path)
    
    print("\n" + "=" * 60)
    print("📊 UPH分析结果")
    print("=" * 60)
    print(f"  LaserIndex数量: {len(data_by_laser)}")
    print(f"  产品总数: {total_products}")
    print(f"  正常产品数: {total_normal_count}")
    print(f"  异常产品数: {total_abnormal_count}")
    print(f"  异常率: {round(total_abnormal_count/total_products*100, 2) if total_products > 0 else 0}%")
    print("-" * 60)
    print(f"  正常产品UPH预估: {normal_uph}")
    print(f"  实际UPH: {actual_uph}")
    print("=" * 60)
    print(f"✅ 转换完成！Excel文件已保存为：{output_file_path}")
    print("   📌 所有公式已清除，Numbers可正常打开")
    
    return output_file_path


# ... 前面是所有函数定义 (clean_value_for_excel, parse_time_to_datetime, read_data_from_file, calculate_uph 等) ...

# ============================================================
# GUI文件选择函数
# ============================================================

def select_file():
    root = tk.Tk()
    root.withdraw()  # 隐藏主窗口
    file_path = filedialog.askopenfilename(
        title="请选择CSV或Excel日志文件",
        filetypes=[("CSV文件", "*.csv"), ("Excel文件", "*.xlsx")]
    )
    return file_path


# ============================================================
# 程序入口
# ============================================================
if __name__ == "__main__":
    file_path = select_file()
    if file_path:
        calculate_uph(file_path)
        print("处理完成！")
        input("按Enter键退出...")
    else:
        print("未选择文件，程序退出")
        input("按Enter键退出...")